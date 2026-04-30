#!/usr/bin/env python3
"""Convert legacy Excel files (.xls) to modern .xlsx format.

Primary mode  : Excel COM automation (Windows + Excel required).
                Opens each file exactly as Excel would and saves it as .xlsx —
                equivalent to File → Save As, preserving all formatting.
Fallback mode : Pure-Python parser (xlrd / SpreadsheetML / HTML).
                Data values only, no formatting.  Activated automatically when
                Excel is unavailable, or explicitly with --no-excel.

Author:  GFrancV
Version: 0.2.0
"""

__version__ = "0.2.0"
__author__ = "GFrancV"

import argparse
import os
import re
import sys
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path

# Ensure UTF-8 output on Windows (avoids garbled chars in help/progress)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

import xlrd
import openpyxl

OLD_FORMATS = {".xls"}

# OLE2 Compound Document magic bytes — present in every genuine .xls binary file
_OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0"

# SpreadsheetML XML namespace
_SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"

# Pre-built namespace-qualified tag/attribute names used in SpreadsheetML parsing
_SS = {
    "Worksheet": f"{{{_SS_NS}}}Worksheet",
    "Table":     f"{{{_SS_NS}}}Table",
    "Row":       f"{{{_SS_NS}}}Row",
    "Cell":      f"{{{_SS_NS}}}Cell",
    "Data":      f"{{{_SS_NS}}}Data",
    "Name":      f"{{{_SS_NS}}}Name",
    "Index":     f"{{{_SS_NS}}}Index",
    "Merge":     f"{{{_SS_NS}}}MergeAcross",
    "Type":      f"{{{_SS_NS}}}Type",
}

# Excel SaveAs format constant for .xlsx (no macros)
_XLSX_FORMAT = 51  # xlOpenXMLWorkbook


# ══════════════════════════════════════════════════════════════════════════════
#  COM-BASED CONVERSION  (primary mode — requires Windows + Excel)
# ══════════════════════════════════════════════════════════════════════════════

def _start_excel():
    """
    Launch a hidden Excel instance via COM.
    Returns the Application object, or None if COM / Excel is unavailable.
    """
    try:
        import pythoncom
        import win32com.client
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
        return excel
    except Exception:
        return None


def _stop_excel(excel) -> None:
    """Quit an Excel COM instance and release the COM apartment."""
    try:
        excel.Quit()
    except Exception:
        pass
    try:
        import pythoncom
        pythoncom.CoUninitialize()
    except Exception:
        pass


def _convert_with_excel(excel, src: Path, dst: Path) -> None:
    """
    Open *src* in the provided Excel instance and save as .xlsx to *dst*.

    Handles automatically:
    - Format-mismatch dialogs (DisplayAlerts = False)
    - Protected View (files with Zone.Identifier / internet-origin mark)
    """
    wb = None
    try:
        dst.parent.mkdir(parents=True, exist_ok=True)
        pv_before = excel.ProtectedViewWindows.Count

        wb = excel.Workbooks.Open(
            Filename=str(src.resolve()),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
            Notify=False,
        )

        # If the file was intercepted by Protected View, exit it to get
        # a writable workbook reference before calling SaveAs.
        if excel.ProtectedViewWindows.Count > pv_before:
            pv = excel.ProtectedViewWindows.Item(excel.ProtectedViewWindows.Count)
            wb = pv.Edit()

        wb.SaveAs(
            Filename=str(dst.resolve()),
            FileFormat=_XLSX_FORMAT,
            CreateBackup=False,
        )
        wb.Close(SaveChanges=False)
        wb = None

    except Exception:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        for _ in range(excel.ProtectedViewWindows.Count):
            try:
                excel.ProtectedViewWindows.Item(1).Close()
            except Exception:
                break
        raise


# ══════════════════════════════════════════════════════════════════════════════
#  PURE-PYTHON FALLBACK  (data-only, no Excel required)
# ══════════════════════════════════════════════════════════════════════════════

def _detect_format(path: Path) -> str:
    """
    Read the first 1024 bytes to determine the actual format of a .xls file:
      'xls'  — OLE2 binary (Excel 97-2003 binary format)
      'xml'  — SpreadsheetML (XML-based, Excel 2002-2003)
      'html' — HTML table saved with .xls extension
    """
    with open(path, "rb") as f:
        head = f.read(1024)
    if head[:4] == _OLE2_SIGNATURE:
        return "xls"
    if b"<?xml" in head or b"<Workbook" in head or b"schemas-microsoft" in head:
        return "xml"
    return "html"


class _TableParser(HTMLParser):
    """Minimal HTML <table> extractor for the HTML fallback path."""

    def __init__(self):
        super().__init__()
        self.tables: list = []
        self._table = None
        self._row = None
        self._buf = None

    def handle_starttag(self, tag, attrs):
        t = tag.lower()
        if t == "table":
            self._table = []
        elif t == "tr" and self._table is not None:
            self._row = []
        elif t in ("td", "th") and self._row is not None:
            self._buf = []

    def handle_endtag(self, tag):
        t = tag.lower()
        if t == "table" and self._table is not None:
            if self._table:
                self.tables.append(self._table)
            self._table = None
        elif t == "tr" and self._table is not None and self._row is not None:
            if self._row:
                self._table.append(self._row)
            self._row = None
        elif t in ("td", "th") and self._buf is not None and self._row is not None:
            self._row.append("".join(self._buf).strip())
            self._buf = None

    def handle_data(self, data):
        if self._buf is not None:
            self._buf.append(data)


def _coerce(raw: str):
    """Convert a string cell value to int or float if possible; keep as str otherwise."""
    if not raw:
        return None
    try:
        return int(raw)
    except ValueError:
        pass
    try:
        return float(raw)
    except ValueError:
        pass
    return raw


def _sheets_from_html(path: Path) -> list:
    """Parse an HTML-disguised .xls file. Returns [(sheet_name, rows), ...]."""
    raw = path.read_bytes()
    m = re.search(rb'charset=["\']?([\w-]+)', raw[:2048], re.IGNORECASE)
    encoding = m.group(1).decode("ascii") if m else "utf-8"
    try:
        html = raw.decode(encoding, errors="replace")
    except LookupError:
        html = raw.decode("latin-1", errors="replace")
    parser = _TableParser()
    parser.feed(html)
    return [
        (f"Sheet{i + 1}", [[_coerce(v) for v in row] for row in table])
        for i, table in enumerate(parser.tables)
    ]


def _sheets_from_xml(path: Path) -> list:
    """
    Parse a SpreadsheetML .xls file. Returns [(sheet_name, rows), ...].

    Handles pre-XML comments, sparse columns (ss:Index), merged-column
    tracking (ss:MergeAcross), and XML 1.0 illegal control characters.
    """
    raw = path.read_bytes()
    for marker in (b"<?xml", b"<Workbook"):
        idx = raw.find(marker)
        if idx > 0:
            raw = raw[idx:]
            break
    # Strip control characters disallowed by XML 1.0 (except tab \x09, LF \x0A, CR \x0D)
    raw = re.sub(rb"[\x00-\x08\x0b\x0c\x0e-\x1f]", b"", raw)
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        # Last resort: decode as latin-1 and strip the encoding declaration so
        # Python's internal UTF-16 handling can take over.
        text = raw.decode("latin-1", errors="replace")
        text = re.sub(r"<\?xml[^?]*\?>", "", text, count=1)
        root = ET.fromstring(text)

    sheets = []
    for ws_el in root.iter(_SS["Worksheet"]):
        name = ws_el.get(_SS["Name"], f"Sheet{len(sheets) + 1}")
        table_el = ws_el.find(_SS["Table"])
        if table_el is None:
            continue
        rows = []
        for row_el in table_el.findall(_SS["Row"]):
            row_data: dict = {}
            col = 1
            for cell_el in row_el.findall(_SS["Cell"]):
                idx_attr = cell_el.get(_SS["Index"])
                if idx_attr:
                    col = int(idx_attr)
                merge = cell_el.get(_SS["Merge"])
                data_el = cell_el.find(_SS["Data"])
                if data_el is not None and data_el.text:
                    dtype = data_el.get(_SS["Type"], "String")
                    raw_val = data_el.text
                    if dtype == "Number":
                        try:
                            v = int(raw_val) if "." not in raw_val else float(raw_val)
                        except ValueError:
                            v = raw_val
                    elif dtype == "Boolean":
                        v = raw_val == "1"
                    elif dtype == "DateTime":
                        try:
                            v = datetime.fromisoformat(raw_val.rstrip("Z"))
                        except ValueError:
                            v = raw_val
                    else:
                        v = raw_val or None
                    if v is not None:
                        row_data[col - 1] = v
                col += 1 + (int(merge) if merge else 0)
            if row_data:
                max_col = max(row_data.keys()) + 1
                rows.append([row_data.get(c) for c in range(max_col)])
        if rows:
            sheets.append((name, rows))
    return sheets


def _sheets_from_xls(path: Path) -> list:
    """Read a genuine OLE2 .xls binary file. Returns [(sheet_name, rows), ...]."""
    wb = xlrd.open_workbook(str(path))
    result = []
    for idx in range(wb.nsheets):
        ws = wb.sheet_by_index(idx)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                ct = cell.ctype
                if ct == xlrd.XL_CELL_DATE:
                    v = xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode)
                elif ct == xlrd.XL_CELL_BOOLEAN:
                    v = bool(cell.value)
                elif ct in (xlrd.XL_CELL_ERROR, xlrd.XL_CELL_EMPTY):
                    v = None
                else:
                    v = cell.value
                row.append(v)
            if any(v is not None for v in row):
                rows.append(row)
        result.append((ws.name, rows))
    return result


def _convert_data_only(src: Path, dst: Path) -> tuple:
    """
    Fallback converter: data values only, no formatting.
    Returns (src, ok, error_msg, fmt_tag).
    """
    try:
        dst.parent.mkdir(parents=True, exist_ok=True)
        fmt = _detect_format(src)
        if fmt == "xls":
            sheets = _sheets_from_xls(src)
        elif fmt == "xml":
            sheets = _sheets_from_xml(src)
        else:
            sheets = _sheets_from_html(src)
        if not sheets:
            raise ValueError("No sheets found in file.")
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        for name, rows in sheets:
            ws = wb_out.create_sheet(title=name)
            for row in rows:
                ws.append(row)
        wb_out.save(str(dst))
        return src, True, "", fmt
    except Exception as exc:
        return src, False, str(exc), ""


# ══════════════════════════════════════════════════════════════════════════════
#  FILE DISCOVERY
# ══════════════════════════════════════════════════════════════════════════════

def find_files(input_dir: Path, recursive: bool) -> list:
    """Return all .xls files in input_dir."""
    glob = input_dir.rglob if recursive else input_dir.glob
    return [f for f in glob("*") if f.suffix.lower() in OLD_FORMATS]


def build_tasks(files: list, input_dir: Path, output_dir: Path) -> list:
    """Map each source file to its destination path."""
    return [
        (src, output_dir / src.relative_to(input_dir).with_suffix(".xlsx"))
        for src in files
    ]


# ══════════════════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description=f"Excel Legacy Converter v{__version__} by {__author__} — "
                    "convert .xls files (Excel 97-2003) to modern .xlsx format.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python convert_excel.py ./files\n"
            "  python convert_excel.py ./files ./output\n"
            "  python convert_excel.py ./files --recursive\n"
            "  python convert_excel.py ./files --no-excel   # data-only fallback\n"
        ),
    )
    parser.add_argument("input_dir", help="Folder containing .xls files to convert")
    parser.add_argument(
        "output_dir",
        nargs="?",
        default=None,
        help="Destination folder (default: <input_dir>/converted)",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip Excel COM automation; convert data values only (no formatting)",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=os.cpu_count() or 4,
        metavar="N",
        help="Parallel threads for fallback mode (default: CPU count)",
    )
    parser.add_argument(
        "--recursive",
        "-r",
        action="store_true",
        help="Search subdirectories recursively",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {__version__}",
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir).resolve()
    if not input_dir.is_dir():
        sys.exit(f"Error: '{input_dir}' is not a valid directory.")

    output_dir = (
        Path(args.output_dir).resolve() if args.output_dir else input_dir / "converted"
    )

    files = find_files(input_dir, args.recursive)
    if not files:
        print("No .xls files found in the specified path.")
        return

    tasks = build_tasks(files, input_dir, output_dir)
    ok_count = 0
    failed = []

    # ── COM mode (primary) ────────────────────────────────────────────────────
    excel = None if args.no_excel else _start_excel()

    if excel is not None:
        print(f"Found {len(files)} file(s) to convert...  [Excel COM — full formatting]")
        print(f"Output: {output_dir}\n")
        try:
            for src, dst in tasks:
                try:
                    _convert_with_excel(excel, src, dst)
                    ok_count += 1
                    print(f"  [OK]   {src.name}")
                except Exception as exc:
                    failed.append((src, str(exc)))
                    print(f"  [FAIL] {src.name}: {exc}")
        finally:
            _stop_excel(excel)

    # ── Fallback mode ─────────────────────────────────────────────────────────
    else:
        if not args.no_excel:
            print("Note: Excel COM unavailable — using data-only fallback converter.")
        print(f"Found {len(files)} file(s) to convert...  [data only — no formatting]")
        print(f"Output: {output_dir}\n")

        with ThreadPoolExecutor(max_workers=args.workers) as executor:
            futures = {
                executor.submit(_convert_data_only, src, dst): src
                for src, dst in tasks
            }
            for future in as_completed(futures):
                src, success, error, fmt = future.result()
                if success:
                    ok_count += 1
                    tag = f"  [{fmt.upper()}]" if fmt in ("html", "xml") else ""
                    print(f"  [OK]   {src.name}{tag}")
                else:
                    failed.append((src, error))
                    print(f"  [FAIL] {src.name}: {error}")

    print(f"\nDone: {ok_count} converted, {len(failed)} failed.")
    if failed:
        sys.exit(1)


if __name__ == "__main__":
    main()
