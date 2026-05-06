import re
import xml.etree.ElementTree as ET
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path

import xlrd
import openpyxl

# OLE2 Compound Document magic bytes — present in every genuine .xls binary file
_OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0"

# SpreadsheetML XML namespace
_SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"

# Pre-built namespace-qualified tag/attribute names used in SpreadsheetML parsing
_SS = {
    "Worksheet": f"{{{_SS_NS}}}Worksheet",
    "Table":     f"{{{_SS_NS}}}Table",
    "Row":       f"{{{_SS_NS}}}Row",
    "Cell":      f"{{{_SS_NS}}}Cell",
    "Data":      f"{{{_SS_NS}}}Data",
    "Name":      f"{{{_SS_NS}}}Name",
    "Index":     f"{{{_SS_NS}}}Index",
    "Merge":     f"{{{_SS_NS}}}MergeAcross",
    "Type":      f"{{{_SS_NS}}}Type",
}


def _detect_format(path: Path) -> str:
    """
    Read the first 1024 bytes to determine the actual format of a .xls file:
      'xls'  — OLE2 binary (Excel 97-2003 binary format)
      'xml'  — SpreadsheetML (XML-based, Excel 2002-2003)
      'html' — HTML table saved with .xls extension
    """
    with open(path, "rb") as f:
        head = f.read(1024)
    if head[:4] == _OLE2_SIGNATURE:
        return "xls"
    if b"<?xml" in head or b"<Workbook" in head or b"schemas-microsoft" in head:
        return "xml"
    return "html"


class _TableParser(HTMLParser):
    """Minimal HTML <table> extractor for the HTML fallback path."""

    def __init__(self):
        super().__init__()
        self.tables: list = []
        self._table = None
        self._row = None
        self._buf = None

    def handle_starttag(self, tag, attrs):
        t = tag.lower()
        if t == "table":
            self._table = []
        elif t == "tr" and self._table is not None:
            self._row = []
        elif t in ("td", "th") and self._row is not None:
            self._buf = []

    def handle_endtag(self, tag):
        t = tag.lower()
        if t == "table" and self._table is not None:
            if self._table:
                self.tables.append(self._table)
            self._table = None
        elif t == "tr" and self._table is not None and self._row is not None:
            if self._row:
                self._table.append(self._row)
            self._row = None
        elif t in ("td", "th") and self._buf is not None and self._row is not None:
            self._row.append("".join(self._buf).strip())
            self._buf = None

    def handle_data(self, data):
        if self._buf is not None:
            self._buf.append(data)


def _coerce(raw: str):
    """Convert a string cell value to int or float if possible; keep as str otherwise."""
    if not raw:
        return None
    try:
        return int(raw)
    except ValueError:
        pass
    try:
        return float(raw)
    except ValueError:
        pass
    return raw


def _sheets_from_html(path: Path) -> list:
    """Parse an HTML-disguised .xls file. Returns [(sheet_name, rows), ...]."""
    raw = path.read_bytes()
    m = re.search(rb'charset=["\']?([\w-]+)', raw[:2048], re.IGNORECASE)
    encoding = m.group(1).decode("ascii") if m else "utf-8"
    try:
        html = raw.decode(encoding, errors="replace")
    except LookupError:
        html = raw.decode("latin-1", errors="replace")
    parser = _TableParser()
    parser.feed(html)
    return [
        (f"Sheet{i + 1}", [[_coerce(v) for v in row] for row in table])
        for i, table in enumerate(parser.tables)
    ]


def _sheets_from_xml(path: Path) -> list:
    """
    Parse a SpreadsheetML .xls file. Returns [(sheet_name, rows), ...].

    Handles pre-XML comments, sparse columns (ss:Index), merged-column
    tracking (ss:MergeAcross), and XML 1.0 illegal control characters.
    """
    raw = path.read_bytes()
    for marker in (b"<?xml", b"<Workbook"):
        idx = raw.find(marker)
        if idx > 0:
            raw = raw[idx:]
            break
    # Strip control characters disallowed by XML 1.0 (except tab \x09, LF \x0A, CR \x0D)
    raw = re.sub(rb"[\x00-\x08\x0b\x0c\x0e-\x1f]", b"", raw)
    try:
        root = ET.fromstring(raw)
    except ET.ParseError:
        # Last resort: decode as latin-1 and strip the encoding declaration so
        # Python's internal UTF-16 handling can take over.
        text = raw.decode("latin-1", errors="replace")
        text = re.sub(r"<\?xml[^?]*\?>", "", text, count=1)
        root = ET.fromstring(text)

    sheets = []
    for ws_el in root.iter(_SS["Worksheet"]):
        name = ws_el.get(_SS["Name"], f"Sheet{len(sheets) + 1}")
        table_el = ws_el.find(_SS["Table"])
        if table_el is None:
            continue
        rows = []
        for row_el in table_el.findall(_SS["Row"]):
            row_data: dict = {}
            col = 1
            for cell_el in row_el.findall(_SS["Cell"]):
                idx_attr = cell_el.get(_SS["Index"])
                if idx_attr:
                    col = int(idx_attr)
                merge = cell_el.get(_SS["Merge"])
                data_el = cell_el.find(_SS["Data"])
                if data_el is not None and data_el.text:
                    dtype = data_el.get(_SS["Type"], "String")
                    raw_val = data_el.text
                    if dtype == "Number":
                        try:
                            v = int(raw_val) if "." not in raw_val else float(raw_val)
                        except ValueError:
                            v = raw_val
                    elif dtype == "Boolean":
                        v = raw_val == "1"
                    elif dtype == "DateTime":
                        try:
                            v = datetime.fromisoformat(raw_val.rstrip("Z"))
                        except ValueError:
                            v = raw_val
                    else:
                        v = raw_val or None
                    if v is not None:
                        row_data[col - 1] = v
                col += 1 + (int(merge) if merge else 0)
            if row_data:
                max_col = max(row_data.keys()) + 1
                rows.append([row_data.get(c) for c in range(max_col)])
        if rows:
            sheets.append((name, rows))
    return sheets


def _sheets_from_xls(path: Path) -> list:
    """Read a genuine OLE2 .xls binary file. Returns [(sheet_name, rows), ...]."""
    wb = xlrd.open_workbook(str(path))
    result = []
    for idx in range(wb.nsheets):
        ws = wb.sheet_by_index(idx)
        rows = []
        for r in range(ws.nrows):
            row = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                ct = cell.ctype
                if ct == xlrd.XL_CELL_DATE:
                    v = xlrd.xldate.xldate_as_datetime(cell.value, wb.datemode)
                elif ct == xlrd.XL_CELL_BOOLEAN:
                    v = bool(cell.value)
                elif ct in (xlrd.XL_CELL_ERROR, xlrd.XL_CELL_EMPTY):
                    v = None
                else:
                    v = cell.value
                row.append(v)
            if any(v is not None for v in row):
                rows.append(row)
        result.append((ws.name, rows))
    return result


def _convert_data_only(src: Path, dst: Path) -> tuple:
    """
    Fallback converter: data values only, no formatting.
    Returns (src, ok, error_msg, fmt_tag).
    """
    try:
        dst.parent.mkdir(parents=True, exist_ok=True)
        fmt = _detect_format(src)
        if fmt == "xls":
            sheets = _sheets_from_xls(src)
        elif fmt == "xml":
            sheets = _sheets_from_xml(src)
        else:
            sheets = _sheets_from_html(src)
        if not sheets:
            raise ValueError("No sheets found in file.")
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)
        for name, rows in sheets:
            ws = wb_out.create_sheet(title=name)
            for row in rows:
                ws.append(row)
        wb_out.save(str(dst))
        return src, True, "", fmt
    except Exception as exc:
        return src, False, str(exc), ""
